
from selenium import webdriver
import openpyxl
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import *
import os
import time
from datetime import datetime
from time import sleep
import random
import glob
import schedule

print("               =================================================================================")
print("               ======          AUTOMAÇÃO DO SITE OLX                                     =======")
print("               =================================================================================")
print("               =================================================================================")
print("               ====== BUSCADOR DE APARTAMENTO DA CDHU NA PROMOÇÃO                        =======")

print(os.linesep)

print(os.linesep)

def Buscador_De_Ap_Cdhu():

    print(f'Aqui começar as Configurações do body da Automação....Chrome Options....{os.linesep}')
    Chrome_options = Options()
    Chrome_options.add_argument('--lang=pt-BR')
    Chrome_options.add_argument('--disable-notifications')
    Chrome_options.add_argument('ignore-certificate-errors')
    Chrome_options.add_argument('--ignore-ssl-errors')
    Chrome_options.add_argument("--ignore-certificate-errors")
    Chrome_options.add_argument('--ignore-page_load_metrics_update_dispatcher')
    Chrome_options.add_argument('--disable-gpu')
    Chrome_options.binary_location = os.environ.get('GOOGLE_CHROME_BIN')

    print(os.linesep)
    print(f'Aqui começar as configuração do WebDriver  da Automação....Webdriver......{os.linesep}')
    #  CONFIGURAÇÃO PARA RODAR NO HEROKU
    # RODAR EM SEGUNDO PLANO
    Chrome_options.add_argument('--headless')
    # RODAR SEM ERRO DE POUCA MEMORIA
    Chrome_options.add_argument('--disable-dev-shm-usage')
    # RODAR EM SERVIDOR LINUX
    Chrome_options.add_argument('--no-sandbox')
    caminho_do_Chome_driver = os.environ.get('CHROMEDRIVER_PATH')
    Webdriver = webdriver.Chrome(executable_path=caminho_do_Chome_driver, options=Chrome_options)
    # Webdriver = webdriver.Chrome(executable_path=r'./chromedriver.exe')

    wait = WebDriverWait(
        driver=Webdriver,
        timeout=10,
        poll_frequency=1,
        ignored_exceptions=[NoSuchElementException,
                            ElementNotVisibleException,
                            ElementNotSelectableException]
    )

    print(os.linesep)
    print(f'Aqui começar as configuração Na Página do Site Olx  da Automação....Webdriver......{os.linesep}')
    Webdriver.get("https://sp.olx.com.br/?q=apartamento%20cdhu")
    # SÓ AP DE GUARULHOS
    # Webdriver.get("https://sp.olx.com.br/sao-paulo-e-regiao/outras-cidades/guarulhos/imoveis/venda?q=apartamento%20cdhu")

    Webdriver.maximize_window()
    Hora_Atual = str(datetime.now().strftime('%H:%M'))
    Ano_Atual = str(datetime.now().strftime('%d-%m-%Y'))


    print(os.linesep)
    print(f'Aqui começar as configuração para desabiliatr os Cookis da Página do Site Olx  da Automação....Webdriver......{os.linesep}')
    try:

        cookies_desabilitado = wait.until(
            expected_conditions.presence_of_element_located(
                (By.XPATH, '//button[@id="cookie-notice-ok-button"]')
            )
        )
        if cookies_desabilitado is not None:
            print(f'🙌 🍪 Encontramos a Notificação para Desabilitar o Cookies 🍪 🍪 🍪 {os.linesep}.....Aguarde Por favor')
            Webdriver.execute_script('arguments[0].click()', cookies_desabilitado)
            print('Desabilitamos com Sucesso !!!!!')
    except:
        print('🤔 Não Formos Capaz de Encontra a Notificação dos cookies...')
        print(ErrorInResponseException)
        pass

    print('============================ AQUI COMEÇA A CRIAÇÃO DA PLANILHA EXCEL ============================')
    print(os.linesep)

    print(f'Aqui começar as configuração para Criar Planilha Excel   da Automação.......{os.linesep}')
    criando_planilha = openpyxl.Workbook()
    criando_planilha.create_sheet('Apartamento_Cdhu')
    planilha_apartamento = criando_planilha['Apartamento_Cdhu']
    planilha_apartamento.cell(row=1, column=1, value='Títulos')
    planilha_apartamento.cell(row=1, column=2, value='Preços')
    planilha_apartamento.cell(row=1, column=3, value='Localização')
    planilha_apartamento.cell(row=2,  column=1, value=' ')
    planilha_apartamento.cell(row=2,  column=2, value= ' ')
    planilha_apartamento.cell(row=2,  column=2, value= ' ')

    # planilha_apartamento['A1'] = 'Títulos'
    # planilha_apartamento['B1'] = 'Preços'
    # planilha_apartamento['C1'] = 'Localização'
    # planilha_apartamento['A2'] = ' '
    # planilha_apartamento['B2'] = ' '
    # planilha_apartamento['C2'] = ' '
    print('============================ AQUI FINALIZA A CRIAÇÃO DA PLANILHA EXCEL ============================')

    print(os.linesep)
    print(f'Aqui começar as configurações para Encontra os elementos  da Página do Site Olx  da Automação......{os.linesep}')

    try:

        for i in range(1,23):

            print(os.linesep)
            titulo = wait.until(
                expected_conditions.presence_of_all_elements_located(
                    (By.XPATH, '//div[@class="fnmrjs-6 iNpuEh"]')
                )
            )
            if titulo is not None:
                print(f'🤗 Encontramos Todos os Títulos da Página 🤗 {os.linesep}')

            preco = wait.until(
                expected_conditions.presence_of_all_elements_located(
                    (By.XPATH, '//div[@class="aoie8y-0 hRScWw"]')
                )
            )
            if preco is not None:
                print(
                    f'🙌 Encontramos Todos os preços 🤑 da sua Pesquisa{os.linesep} ')

            localizacao = wait.until(
                expected_conditions.presence_of_all_elements_located(
                    (By.XPATH, '//span[@class="sc-7l84qu-1 ciykCV sc-ifAKCX dpURtf"]')
                )
            )
            if localizacao is not None:
                print(f'🙌 Encontramos Todas as suas Localização da sua Pesquisa{os.linesep}')

            print(os.linesep)
#  '''===========================================================================////////////////////////='''
            print(f'Aqui começar as configurações para Tirar o Print  da Página do Site Olx  da Automação......{os.linesep}')
            Webdriver.execute_script('window.scrollBy(0,800)')
            # VAMOS DESCER 800PIXEL DA PÁGINA PARA PODER TIRA O PRINT
            sleep(random.randint(1,2))
            print(f'🙌 Estamos tirando um Print das Informações do Site 🙌 .....{os.linesep}.....Aguarde alguns instante!')
            print('Eh hora de tira um Print da Página !!!!!')
            time.sleep(random.randint(2, 3))
            nome_arquivo = str(datetime.now().strftime('%d-%m-%Y -%S')) + ".png"
            nome_Arquivo_com_Diretorio = os.path.join('Diretório', nome_arquivo )
            Webdriver.save_screenshot(nome_Arquivo_com_Diretorio)

          
#  :'''===================================================================////////////////////////='''
            # self.webdriver.execute_script('window.scrollBy(0,900)')
            print(os.linesep)
            print('============================ AQUI COMEÇA A INSERÇÃO DAS INFORMAÇÕES DENTRO DA PLANILHA EXCEL ============================')
            print(os.linesep)
            print(f'Aqui começar as configuração para Inseri as Informações dentro do Excel \Automação\......{os.linesep}')

            for indice in range(1, 49):

                nova_linha = [titulo[indice].text,preco[indice].text, localizacao[indice].text]

                planilha_apartamento.append(nova_linha)
                # self.localizacao[informacoes].text]
                print(f'📋 Estamos salvando a sua Pesquisa dentro do Excel 📝📝📝....{indice}')

                criando_planilha.save('Apartamento_CDHU.xlsx')

            print('============================ AQUI FINALIZA A INSERÇÃO DAS INFORMAÇÕES DENTRO DA PLANILHA EXCEL ============================')

            print(os.linesep)
            print(f'Aqui começar as configuração para Encontra a Próxima Página do Site Olx  Automação......{os.linesep}')
            # self.webdriver.execute_script('Window.scrollTo(0,document.body.scrollHeight);') #desce a Página até o Final.

            Webdriver.execute_script('window.scrollTo(0,document.body.scrollHeight);')
            proxima_pagina = wait.until(
                expected_conditions.presence_of_element_located(
                    # (By.XPATH,' //*[starts-with(text(),"Próxima pagina")]')
                    (By.XPATH, '//*[contains(text(),"Próxima pagina")]')
                )
            )
            if proxima_pagina is not None:

                print(os.linesep)
                print(f' ⏭  Encontramos a Página de Próximo ⏭ {os.linesep}')
                sleep(random.randint(2,3))
                Webdriver.execute_script("arguments[0].click()", proxima_pagina)
                print('💯 Acabou de Chegar no Final da Página......💯💯💯')
                print(f'🤖🤖Obrigado por usar o Nosso Boot🤖🤖🤖{os.linesep}')

                Mostrando_o_horario_que_enviou = datetime.now().strftime('%d%m%Y %H:%M')
                mostar_a_data_do_ano = datetime.now().strftime('%d-%m-%Y')

                print(f'🤖🤖Obrigado por usar o Nosso Boot🤖🤖🤖 as {Mostrando_o_horario_que_enviou[9:]} do Dia {mostar_a_data_do_ano}{os.linesep}')
                print('Serviço de Raspagem de Dados terminado com sucesso web screll')
                print(f'🙌 Chegamos ao Final de Todas as Páginas..... até mais 🙌!! as:{Mostrando_o_horario_que_enviou[9:]} do Dia {mostar_a_data_do_ano}')
                time.sleep(3)

    except:
        print('🤔 Não Formos Capaz de Encontra o Erro ..... Não Sei onde está o Erro ')

    from enviando_email_Rodando_todos_os_dias import EnvioDeEmails
    print(
        f'Aqui COMERÇA as configuração para ser enviada  por E-mail as Informações da .... \Automação\......{os.linesep}')
    send_Email = EnvioDeEmails()
    print('Estamos Enviando Seu Email !!!')
    print('Estamos Anexando o arquivo dentro do seu E-mail')
    send_Email.Start_Send()
    print('E-mail Enviado com Sucesso !!!!')

    print(f'Aqui Termina as configuração para ser enviada  por E-mail as Informações da .... \Automação\......{os.linesep}')

    print('# BUSCAR POR DIRETÓRIO.... VAI PROCURA PASTA QUE SE ENCONTRA AS FOTOS OU OS PRINTS ".PNG"')
    targetPatter = os.path.join(os.getcwd() + os.sep + 'Diretório' + os.sep + '*.png')
    # VAMOS ATRIBUIR A UMA VARIAVÉL, PARA BUSCAR TODAS AS INFORMAÇÕES DO targetPatter
    caminho_do_diretorio = glob.glob(targetPatter)
    print(caminho_do_diretorio)
    print(os.linesep)
    print(os.linesep)


    sleep(random.randint(10,15))
    print(f'⏭  Vamos Excluir todas as Fotos com final .png{os.linesep}.....Aguarde{os.linesep}')

    for caminhos_dos_diretorios in caminho_do_diretorio:
        print(f'⏭  Vamos criar um Laço de Repetição  para poder resolver a questão da exclusão em Massa {os.linesep}')
        os.remove(caminhos_dos_diretorios)
        print(f'⏭  Excluimos com Sucesso {os.linesep}')

        Mostrando_o_horario_que_enviou = datetime.now().strftime('%d-%m-%Y')
        mostar_a_data_do_ano = datetime.now().strftime('%d-%m-%Y')
        print(f'💯💯💯 Exclusão feitas as {Mostrando_o_horario_que_enviou} {mostar_a_data_do_ano}{os.linesep}')
        print(f' 🤖🤖 Obrigado por usar o Nosso Boot🤖🤖🤖 até mais...{os.linesep}{os.linesep}')
        print(os.linesep)


schedule.every().days.at('07:27:45').do(Buscador_De_Ap_Cdhu)

while True:
    schedule.run_pending()
    sleep(random.randint(10, 15))
